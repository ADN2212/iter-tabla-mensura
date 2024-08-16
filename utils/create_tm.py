from utils.calcular_rumbo_y_distancia import rumbo_y_dist
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

formato_millares = '#,##0.00'
fuente = Font(name = "Arial", size = 11)
linea_borde = Side(border_style = "thin")
alineacion = Alignment(horizontal = "center")
todos_los_bordes = Border(top = linea_borde, left = linea_borde, right = linea_borde, bottom = linea_borde)

def dar_formato_tabla(hoja, base_row, current_base_clolunm, longitud):
	for fila in hoja.iter_rows(min_row = base_row, max_row = base_row + 1, min_col = current_base_clolunm, max_col = current_base_clolunm + 4):
		for celda in fila:
			celda.font = Font(name = "Arial", size = 12, bold = True)
			celda.border = todos_los_bordes
			celda.alignment = alineacion

	for fila in hoja.iter_rows(min_row = base_row + 2, max_row = base_row + longitud + 1, min_col = current_base_clolunm, max_col = current_base_clolunm + 1):
		for celda in fila:
			celda.font, celda.alignment, celda.border = fuente, alineacion, todos_los_bordes

	for fila in hoja.iter_rows(min_row = base_row + 2, max_row = base_row + longitud + 1, min_col = current_base_clolunm + 2, max_col = current_base_clolunm + 4):
		for celda in fila:
			celda.font, celda.alignment, celda.border, celda.number_format = fuente, alineacion, todos_los_bordes, formato_millares

	hoja.column_dimensions[get_column_letter(current_base_clolunm)].width = 10
	hoja.column_dimensions[get_column_letter(current_base_clolunm + 1)].width = 12
	hoja.column_dimensions[get_column_letter(current_base_clolunm + 2)].width = 12
	hoja.column_dimensions[get_column_letter(current_base_clolunm + 3)].width = 15
	hoja.column_dimensions[get_column_letter(current_base_clolunm + 4)].width = 15
	
	str_casilla = get_column_letter(current_base_clolunm) + str(base_row + longitud + 3)

	hoja[str_casilla].font = Font(size = 12, bold = True, color = "000000FF")
	hoja[str_casilla].alignment = alineacion


def calcular_area(coordenadas):
	area = 0
	for i in range(len(coordenadas) - 1):
		area += coordenadas[i][0] * coordenadas[i + 1][1]
		area -= coordenadas[i][1] * coordenadas[i + 1][0]

	area += coordenadas[len(coordenadas) - 1][0] * coordenadas[0][1] 	
	area -= coordenadas[len(coordenadas) - 1][1] * coordenadas[0][0]

	area = round(0.5*abs(area), 2)

	return area


def create_tables(libro, nombres, set_coordenadas):
	areas = []
	hoja = libro.active
	hoja.title = "Tabla(s)"
	hoja.merge_cells("B1:F1")
	hoja["B1"] = "Por Agrim Juan A. Núñez"
	hoja["B1"].font = Font(name = "times new roman", size = 12, bold = True, color = "000000FF")
	hoja["B1"].alignment = alineacion
	base_row = 3
	current_base_clolunm = 2
	indice_nombre_actual = 0

	for coordenadas_actulaes in set_coordenadas:
		hoja.cell(row = base_row - 1, column = current_base_clolunm, value = f'Para {nombres[indice_nombre_actual]}:')
		hoja.merge_cells(start_row = base_row - 1, start_column = current_base_clolunm, end_row = base_row - 1, end_column = current_base_clolunm + 4)
		hoja.merge_cells(start_row = base_row, start_column = current_base_clolunm, end_row = base_row, end_column = current_base_clolunm + 4)
		hoja.cell(row = base_row, column = current_base_clolunm, value = "Coordenadas UTM Zona 19 Norte")
		hoja.cell(row = base_row + 1, column = current_base_clolunm, value = 'Est.')
		hoja.cell(row = base_row + 1, column = current_base_clolunm + 1, value = 'Rumbo')
		hoja.cell(row = base_row + 1, column = current_base_clolunm + 2, value = 'Dist.(m)')
		hoja.cell(row = base_row + 1, column = current_base_clolunm + 3, value = 'Este(x)')
		hoja.cell(row = base_row + 1, column = current_base_clolunm + 4, value = 'Norte(y)')
		longitud = len(coordenadas_actulaes)
		i = 0

		for fila in hoja.iter_rows(min_row = base_row + 2, max_row = base_row + longitud, min_col = current_base_clolunm, max_col = current_base_clolunm + 4):
			punto_actual = (i + 1, coordenadas_actulaes[i][0], coordenadas_actulaes[i][1])
			punto_siguiente = (None, coordenadas_actulaes[i + 1][0], coordenadas_actulaes[i + 1][1])
			ryd = rumbo_y_dist(punto_actual, punto_siguiente)
			fila[0].value = punto_actual[0]
			fila[1].value = ryd[0]
			fila[2].value = round(ryd[1], 2)
			fila[3].value = round(punto_actual[1], 2)
			fila[4].value = round(punto_actual[2], 2)
			i = i + 1
   
		primer_punto = (None, coordenadas_actulaes[0][0], coordenadas_actulaes[0][1])
		ultimo_punto = (longitud, coordenadas_actulaes[longitud - 1][0], coordenadas_actulaes[longitud - 1][1])
		ryd_final = rumbo_y_dist(ultimo_punto, primer_punto)
		hoja.cell(row = base_row + longitud + 1, column = current_base_clolunm, value = ultimo_punto[0])	
		hoja.cell(row = base_row + longitud + 1, column = current_base_clolunm + 1, value = ryd_final[0])	
		hoja.cell(row = base_row + longitud + 1, column = current_base_clolunm + 2, value = round(ryd_final[1], 2))
		hoja.cell(row = base_row + longitud + 1, column = current_base_clolunm + 3, value = round(ultimo_punto[1], 2))
		hoja.cell(row = base_row + longitud + 1, column = current_base_clolunm + 4, value = round(ultimo_punto[2], 2))
		dar_formato_tabla(hoja, base_row, current_base_clolunm, longitud)
		current_area = calcular_area(coordenadas_actulaes)
		areas.append(current_area)
		area_txt = f'Area = {current_area} m^2' if current_area < 628.86 else f'Area = {current_area} m^2 = {round(current_area/628.86,2)} tareas.'
		hoja.cell(row = base_row + longitud + 3, column = current_base_clolunm, value = area_txt)
		hoja.merge_cells(start_row = base_row + longitud + 3, start_column = current_base_clolunm, end_row = base_row + longitud + 3, end_column = current_base_clolunm + 4)
		current_base_clolunm += 6
		indice_nombre_actual += 1
  
	if len(set_coordenadas) > 1:
		i = 0
		celda_encabezado = hoja.cell(row = base_row, column = current_base_clolunm, value = 'Recuadro de Areas')
		hoja.merge_cells(start_row = base_row, end_row = base_row, start_column = current_base_clolunm, end_column = current_base_clolunm + 1)
		celda_descripcion = hoja.cell(row = base_row + 1, column = current_base_clolunm, value = 'Descripción')
		celda_area = hoja.cell(row = base_row + 1, column = current_base_clolunm + 1, value = 'Area (m^2)')

		for celda in [celda_encabezado, celda_descripcion, celda_area]:
			celda.font = Font(name = "Arial", size = 12, bold = True)
			celda.border = todos_los_bordes
			celda.alignment = alineacion
 
		hoja.cell(row = base_row, column = current_base_clolunm + 1).border = todos_los_bordes
		for fila in hoja.iter_rows(min_row = base_row + 2, max_row = base_row + len(areas) + 1, min_col = current_base_clolunm, max_col = current_base_clolunm + 1):
			fila[0].value = nombres[i]
			fila[0].font = fuente
			fila[0].border = todos_los_bordes
			fila[0].alignment = alineacion
			fila[1].value = areas[i]
			fila[1].font = fuente
			fila[1].border = todos_los_bordes
			fila[1].alignment = alineacion
			fila[1].number_format = formato_millares
			i += 1

		hoja.column_dimensions[get_column_letter(current_base_clolunm)].width = max([len(nombre) for nombre in nombres]) + 5
		hoja.column_dimensions[get_column_letter(current_base_clolunm + 1)].width = 20
		pie_de_tabla = hoja.cell(row = base_row + len(areas) + 2, column = current_base_clolunm, value = 'Total')
		pie_de_tabla.font = Font(name = "Arial", size = 12, bold = True)
		pie_de_tabla.border = todos_los_bordes
		pie_de_tabla.alignment = alineacion		
		total = hoja.cell(row = base_row + len(areas) + 2, column = current_base_clolunm + 1, value = round(sum(areas), 3))
		total.font = Font(name = "Arial", size = 12, bold = True)
		total.border = todos_los_bordes
		total.alignment = alineacion
		total.number_format = formato_millares
